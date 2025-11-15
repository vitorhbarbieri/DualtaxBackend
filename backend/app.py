from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timedelta
import hashlib
import uvicorn
import traceback
import logging
import pandas as pd
import io
import httpx  # Para chamadas HTTP assíncronas

# Import openpyxl para gerar Excel (opcional - só se estiver instalado)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Versão da API para controle de processamento
API_VERSION = "3.4.0"  # Apenas dados reais - erros claros quando API falhar

# -----------------------
# Inicializando FastAPI
# -----------------------
app = FastAPI(
    title="Dualtax API",
    description="MVP Dualtax - Simulação Reforma Tributária",
    version=API_VERSION
)

# -----------------------
# Middleware CORS
# -----------------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -----------------------
# Modelos de dados
# -----------------------
class Nota(BaseModel):
    tipo: str
    valor: float

class FaturamentoMensal(BaseModel):
    mes: str
    mes_formatado: str
    total_entrada: float
    total_saida: float
    quantidade_notas_entrada: int
    quantidade_notas_saida: int

class Resultado(BaseModel):
    nome_empresa: str  # Nome da empresa
    cnpj: str  # CNPJ formatado
    total_entrada: float
    total_saida: float
    impacto_cbs: float
    impacto_ibs: float
    impacto_previdencia: float
    impacto_total: float
    periodo_inicio: str
    periodo_fim: str
    periodo_mes: str
    horizonte_temporal: str
    faturamento_mensal: List[FaturamentoMensal]
    debug_info: Optional[dict] = None  # Informações de debug (versão, colunas identificadas)
    validacao_tributaria: Optional[dict] = None  # Validação do agente tributário

# -----------------------
# Função determinística para buscar notas
# -----------------------
async def buscar_dados_empresa_api(cnpj: str) -> dict:
    """
    Busca dados da empresa via API externa (BrasilAPI).
    Retorna: nome, CNAE, regime tributário, situação, etc.
    NÃO USA FALLBACK - APENAS DADOS REAIS.
    """
    cnpj_limpo = ''.join(filter(str.isdigit, cnpj))
    
    if len(cnpj_limpo) != 14:
        raise ValueError(f"CNPJ inválido: deve conter 14 dígitos. Recebido: {len(cnpj_limpo)} dígitos")
    
    # BrasilAPI (gratuita, oficial)
    url = f"https://brasilapi.com.br/api/cnpj/v1/{cnpj_limpo}"
    
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.get(url)
            
            if response.status_code == 200:
                data = response.json()
                logger.info(f"✅ Dados da empresa obtidos via API: {data.get('nome', 'N/A')}")
                logger.info(f"   Estrutura recebida: {list(data.keys())[:10]}")
                
                # Extrai CNAE principal - BrasilAPI retorna 'cnae_fiscal' (int) e 'cnae_fiscal_descricao' (str)
                cnae_principal = None
                cnae_descricao = 'N/A'
                
                # BrasilAPI usa 'cnae_fiscal' (int) e 'cnae_fiscal_descricao' (str)
                if 'cnae_fiscal' in data:
                    cnae_principal = str(data['cnae_fiscal'])  # Converte int para string
                    cnae_descricao = data.get('cnae_fiscal_descricao', 'N/A')
                
                # Fallback: tenta cnae_fiscal_principal (formato alternativo de outras APIs)
                if not cnae_principal and 'cnae_fiscal_principal' in data:
                    cnae_data = data['cnae_fiscal_principal']
                    if isinstance(cnae_data, dict):
                        cnae_principal = str(cnae_data.get('codigo', ''))
                        cnae_descricao = cnae_data.get('descricao', 'N/A')
                    elif isinstance(cnae_data, (str, int)):
                        cnae_principal = str(cnae_data)
                
                # Fallback: tenta atividade_principal (formato alternativo)
                if not cnae_principal and 'atividade_principal' in data:
                    if isinstance(data['atividade_principal'], list) and len(data['atividade_principal']) > 0:
                        atividade = data['atividade_principal'][0]
                        if isinstance(atividade, dict):
                            cnae_principal = str(atividade.get('code') or atividade.get('codigo', ''))
                            cnae_descricao = atividade.get('text') or atividade.get('descricao', 'N/A')
                
                # Extrai nome da empresa (BrasilAPI usa 'razao_social' ou 'nome')
                nome_empresa = data.get('razao_social') or data.get('nome') or 'N/A'
                
                # Extrai situação cadastral
                situacao = data.get('descricao_situacao_cadastral') or data.get('situacao') or 'N/A'
                
                # Extrai regime tributário (BrasilAPI tem 'regime_tributario' como lista)
                regime = 'N/A'
                if 'regime_tributario' in data:
                    if isinstance(data['regime_tributario'], list) and len(data['regime_tributario']) > 0:
                        regime = ', '.join(data['regime_tributario'])
                    elif isinstance(data['regime_tributario'], str):
                        regime = data['regime_tributario']
                elif 'descricao_porte' in data:
                    regime = data['descricao_porte']
                elif 'porte' in data:
                    regime = data['porte']
                
                logger.info(f"   CNAE extraído: {cnae_principal} - {cnae_descricao}")
                logger.info(f"   Situação: {situacao}")
                
                return {
                    "nome_empresa": nome_empresa,
                    "cnpj_formatado": data.get('cnpj', cnpj_limpo),
                    "cnae_principal": cnae_principal,
                    "cnae_descricao": cnae_descricao,
                    "situacao": situacao,
                    "regime_tributario": regime,
                    "uf": data.get('uf', 'N/A'),
                    "municipio": data.get('municipio', 'N/A'),
                    "fonte": "API Externa"
                }
            elif response.status_code == 404:
                raise HTTPException(
                    status_code=404,
                    detail=f"CNPJ não encontrado na base de dados da Receita Federal. Verifique se o CNPJ está correto: {cnpj_limpo}"
                )
            else:
                error_data = response.text
                logger.error(f"API retornou status {response.status_code}: {error_data}")
                raise HTTPException(
                    status_code=response.status_code,
                    detail=f"Erro ao consultar API externa (BrasilAPI). Status: {response.status_code}. Tente novamente mais tarde."
                )
                
    except httpx.TimeoutException:
        logger.error("Timeout ao consultar API externa")
        raise HTTPException(
            status_code=504,
            detail="Timeout ao consultar API externa. A API pode estar sobrecarregada. Tente novamente em alguns instantes."
        )
    except httpx.RequestError as e:
        logger.error(f"Erro de conexão com API: {str(e)}")
        raise HTTPException(
            status_code=503,
            detail=f"Erro de conexão com a API externa. Verifique sua conexão com a internet. Erro: {str(e)}"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro inesperado ao buscar dados via API: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Erro inesperado ao consultar dados da empresa. Tente novamente mais tarde. Erro: {str(e)}"
        )

async def buscar_notas_fiscais_api(cnpj: str, periodo_inicio: str = None, periodo_fim: str = None) -> List[dict]:
    """
    Busca notas fiscais via API externa (NFe.io, APIs de contabilidade, etc.).
    Retorna lista de notas com detalhes: NCM, alíquotas, valores, etc.
    
    NÃO USA FALLBACK - APENAS DADOS REAIS.
    Se não houver integração implementada, retorna lista vazia.
    
    Exemplo de integração com APIs:
    - NFe.io: https://nfe.io/docs
    - Omie: https://developer.omie.com.br/
    - ContaAzul: https://developers.contaazul.com/
    """
    cnpj_limpo = ''.join(filter(str.isdigit, cnpj))
    
    logger.info(f"🔍 Buscando notas fiscais via API para CNPJ: {cnpj_limpo}")
    
    # TODO: Implementar integração real com API de notas fiscais
    # Por enquanto, retorna lista vazia (sem dados mockados)
    # Quando implementar, descomente e configure:
    # async with httpx.AsyncClient(timeout=30.0) as client:
    #     response = await client.get(
    #         f"https://api.exemplo.com/notas/{cnpj_limpo}",
    #         headers={"Authorization": f"Bearer {API_KEY}"},
    #         params={"periodo_inicio": periodo_inicio, "periodo_fim": periodo_fim}
    #     )
    #     if response.status_code == 200:
    #         return response.json()
    
    logger.info("⚠️ Integração com API de notas fiscais não implementada. Retornando lista vazia (sem dados mockados).")
    return []

# -----------------------
# Função para gerar faturamento mensal (APENAS DADOS REAIS)
# -----------------------
def gerar_faturamento_mensal(cnpj: str) -> List[FaturamentoMensal]:
    """
    Gera faturamento mensal baseado em dados reais.
    NÃO USA DADOS MOCKADOS - retorna lista vazia se não houver dados reais.
    
    TODO: Implementar busca de faturamento real via API de notas fiscais ou planilha.
    """
    logger.info(f"gerar_faturamento_mensal chamado com CNPJ: {cnpj}")
    
    # Por enquanto, retorna lista vazia (sem dados mockados)
    # Quando houver integração com API de notas fiscais, calcular faturamento real
    logger.info("⚠️ Faturamento mensal não disponível. Retornando lista vazia (sem dados mockados).")
    return []

# -----------------------
# Função para calcular período de consulta
# -----------------------
def calcular_periodo() -> dict:
    """Calcula o período de consulta - focado em 2025."""
    hoje = datetime.now()
    ano_consulta = 2025
    
    # Se estamos em 2025, usa até o mês atual, senão usa até dezembro
    if hoje.year == 2025:
        periodo_fim = hoje.strftime("%Y-%m-%d")
        mes_atual = hoje.month
    else:
        periodo_fim = f"{ano_consulta}-12-31"
        mes_atual = 12
    
    meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]
    
    return {
        "periodo_inicio": f"{ano_consulta}-01-01",
        "periodo_fim": periodo_fim,
        "periodo_mes": f"{meses[mes_atual - 1]}/{ano_consulta}",
        "horizonte_temporal": f"Ano {ano_consulta} - Período completo"
    }

# -----------------------
# Agente Especialista em Tributos - Validação de Dados
# -----------------------
class ValidacaoTributaria(BaseModel):
    valido: bool
    mensagem: str
    dados_suficientes: bool
    dados_faltantes: List[str]
    recomendacoes: List[str]
    nivel_confianca: str  # "Alto", "Médio", "Baixo"

def consultar_agente_tributario(df: pd.DataFrame, cnpj: str) -> ValidacaoTributaria:
    """
    Agente especialista em tributos analisa se os dados são suficientes
    para calcular o impacto da reforma tributária.
    """
    logger.info("🔍 Agente Tributário: Analisando dados fornecidos...")
    
    # Normaliza colunas
    df.columns = df.columns.str.strip().str.lower()
    colunas_encontradas = list(df.columns)
    
    # Identifica colunas necessárias
    col_mes = None
    col_entrada = None
    col_saida = None
    col_qtd_entrada = None
    col_qtd_saida = None
    
    dados_faltantes = []
    recomendacoes = []
    
    # Verifica colunas obrigatórias
    for col in df.columns:
        col_lower = col.lower()
        if 'mes' in col_lower or 'mês' in col_lower or 'periodo' in col_lower or 'data' in col_lower:
            col_mes = col
        elif 'entrada' in col_lower or 'receita' in col_lower or 'faturamento' in col_lower:
            col_entrada = col
        elif 'saida' in col_lower or 'saída' in col_lower or 'despesa' in col_lower:
            col_saida = col
        elif 'qtd' in col_lower and 'entrada' in col_lower:
            col_qtd_entrada = col
        elif 'qtd' in col_lower and 'saida' in col_lower:
            col_qtd_saida = col
    
    # Validações do agente
    if not col_mes:
        dados_faltantes.append("Coluna de Mês/Período")
    if not col_entrada:
        dados_faltantes.append("Coluna de Entradas/Receita")
    if not col_saida:
        dados_faltantes.append("Coluna de Saídas/Despesas")
    
    # Análise de qualidade dos dados
    dados_suficientes = len(dados_faltantes) == 0
    
    if dados_suficientes:
        # Verifica se há dados válidos
        try:
            entradas_validas = df[col_entrada].notna().sum()
            saidas_validas = df[col_saida].notna().sum()
            total_linhas = len(df)
            
            if total_linhas == 0:
                dados_suficientes = False
                dados_faltantes.append("Nenhuma linha de dados encontrada")
            elif entradas_validas < total_linhas * 0.8:
                recomendacoes.append("⚠️ Algumas linhas de Entradas estão vazias. Preencha para cálculo mais preciso.")
            elif saidas_validas < total_linhas * 0.8:
                recomendacoes.append("⚠️ Algumas linhas de Saídas estão vazias. Preencha para cálculo mais preciso.")
            
            # Verifica período mínimo
            if total_linhas < 3:
                recomendacoes.append("💡 Recomendamos pelo menos 3 meses de dados para uma análise mais precisa.")
            
            # Verifica valores zero
            total_entrada = df[col_entrada].sum() if col_entrada else 0
            total_saida = df[col_saida].sum() if col_saida else 0
            
            if total_entrada == 0:
                recomendacoes.append("⚠️ Total de entradas é zero. Verifique se os dados estão corretos.")
            if total_saida == 0:
                recomendacoes.append("⚠️ Total de saídas é zero. Verifique se os dados estão corretos.")
                
        except Exception as e:
            logger.warning(f"Erro ao analisar dados: {str(e)}")
            recomendacoes.append("⚠️ Erro ao processar alguns dados. Verifique o formato da planilha.")
    
    # Determina nível de confiança
    if dados_suficientes and len(recomendacoes) == 0:
        nivel_confianca = "Alto"
        mensagem = "✅ Dados suficientes para calcular o impacto da reforma tributária com alta precisão."
    elif dados_suficientes and len(recomendacoes) <= 2:
        nivel_confianca = "Médio"
        mensagem = "✅ Dados suficientes, mas há algumas recomendações para melhorar a precisão."
    elif dados_suficientes:
        nivel_confianca = "Médio"
        mensagem = "⚠️ Dados suficientes, mas há várias recomendações importantes."
    else:
        nivel_confianca = "Baixo"
        mensagem = "❌ Dados insuficientes. Faltam informações essenciais para o cálculo."
    
    # Adiciona recomendações específicas do agente tributário
    if not col_qtd_entrada and not col_qtd_saida:
        recomendacoes.append("💡 Adicionar quantidade de notas ajuda a validar os valores informados.")
    
    recomendacoes.append("📊 Para cálculo preciso de CBS (12%) e IBS (5%), precisamos do total de saídas.")
    recomendacoes.append("📊 Para cálculo de Previdência (2%), precisamos do total de entradas.")
    recomendacoes.append("📅 Certifique-se de que os períodos estão corretos (2025 para análise atual).")
    
    logger.info(f"Agente Tributário: Validação concluída - {mensagem}")
    
    return ValidacaoTributaria(
        valido=dados_suficientes,
        mensagem=mensagem,
        dados_suficientes=dados_suficientes,
        dados_faltantes=dados_faltantes,
        recomendacoes=recomendacoes,
        nivel_confianca=nivel_confianca
    )

# -----------------------
# Função para processar planilha de faturamento
# -----------------------
def processar_planilha_faturamento(df: pd.DataFrame, cnpj: str, nome_empresa: str) -> dict:
    """Processa planilha Excel/CSV com faturamento mensal."""
    logger.info(f"Processando planilha para CNPJ: {cnpj}")
    
    # IMPORTANTE: Guarda colunas ORIGINAIS antes de normalizar
    colunas_originais = {col.lower().strip(): col for col in df.columns}
    
    # Normaliza nomes das colunas (remove espaços, converte para minúsculas)
    df.columns = df.columns.str.strip().str.lower()
    
    # Tenta identificar as colunas (flexível com diferentes nomes)
    col_mes = None
    col_entrada = None
    col_saida = None
    col_qtd_entrada = None
    col_qtd_saida = None
    
    # Busca colunas por nomes comuns
    # ESTRATÉGIA: Primeiro identifica TODAS as colunas, depois escolhe as corretas
    # Template tem: "Mês/Ano", "Entradas (R$)", "Saídas (R$)", "Qtd. Notas Entrada", "Qtd. Notas Saída"
    
    todas_colunas = list(df.columns)
    
    print("\n" + "="*80)
    print("📋 COLUNAS ENCONTRADAS NO EXCEL")
    print("="*80)
    print("Colunas ORIGINAIS (antes normalização):")
    for i, col_orig in enumerate(df.columns, 1):
        col_orig_real = colunas_originais.get(col_orig.lower().strip(), col_orig)
        print(f"  {i}. '{col_orig_real}' → normalizada: '{col_orig}'")
    print("="*80)
    
    # Mostra PRIMEIROS VALORES de cada coluna para debug
    print("\n" + "="*80)
    print("🔍 PRIMEIROS VALORES DE CADA COLUNA (para verificar conteúdo)")
    print("="*80)
    for col in todas_colunas[:10]:  # Primeiras 10 colunas
        if len(df) > 0:
            primeiro_valor = df[col].iloc[0] if len(df) > 0 else "N/A"
            segundo_valor = df[col].iloc[1] if len(df) > 1 else "N/A"
            print(f"  '{col}':")
            print(f"    Linha 1: {repr(primeiro_valor)} (tipo: {type(primeiro_valor).__name__})")
            print(f"    Linha 2: {repr(segundo_valor)} (tipo: {type(segundo_valor).__name__})")
    print("="*80 + "\n")
    
    logger.info(f"Todas as colunas encontradas (antes normalização): {todas_colunas}")
    
    # ESTRATÉGIA CORRIGIDA: Identifica colunas de forma mais precisa
    # Template tem exatamente: "Mês/Ano", "Entradas (R$)", "Saídas (R$)", "Qtd. Notas Entrada", "Qtd. Notas Saída"
    
    print("\n" + "="*80)
    print("🔍 IDENTIFICANDO COLUNAS...")
    print("="*80)
    
    # Primeira passada: identifica colunas de QUANTIDADE (mais específicas - contém "qtd")
    for col in todas_colunas:
        col_lower = col.lower().strip()
        
        # Mês/Período
        if ('mes' in col_lower or 'mês' in col_lower or 'periodo' in col_lower or 'data' in col_lower) and col_mes is None:
            col_mes = col
            print(f"  ✓ Mês identificado: '{col}'")
        
        # Quantidades - busca por "qtd" primeiro (mais específico)
        if 'qtd' in col_lower:
            if 'entrada' in col_lower and col_qtd_entrada is None:
                col_qtd_entrada = col
                print(f"  ✓ Qtd. Entrada identificada: '{col}'")
            elif ('saida' in col_lower or 'saída' in col_lower) and col_qtd_saida is None:
                col_qtd_saida = col
                print(f"  ✓ Qtd. Saída identificada: '{col}'")
    
    # Segunda passada: identifica colunas de VALOR (evita as de quantidade)
    # ESTRATÉGIA: Busca primeiro pelos nomes EXATOS do template, depois fallback
    print(f"\n  Buscando colunas de VALOR (pulando quantidades)...")
    
    # PRIORIDADE 1: Busca pelos nomes EXATOS do template (case-insensitive)
    nomes_exatos_entrada = ['entradas (r$)', 'entradas(r$)', 'entrada (r$)', 'entrada(r$)']
    nomes_exatos_saida = ['saídas (r$)', 'saidas (r$)', 'saídas(r$)', 'saidas(r$)', 'saída (r$)', 'saida (r$)']
    
    for col in todas_colunas:
        col_lower = col.lower().strip()
        
        # PULA se já foi identificada como quantidade
        if col == col_qtd_entrada or col == col_qtd_saida:
            print(f"  ⏭️  Pulando coluna de quantidade: '{col}'")
            continue
        
        # PULA se contém "qtd" (já foi processada como quantidade)
        if 'qtd' in col_lower:
            print(f"  ⏭️  Pulando coluna com 'qtd': '{col}'")
            continue
        
        # PRIORIDADE 1: Busca por nome EXATO do template
        if col_entrada is None and col_lower in nomes_exatos_entrada:
            col_entrada = col
            print(f"  ✅✅ ENTRADA identificada (NOME EXATO): '{col}'")
            continue
        
        if col_saida is None and col_lower in nomes_exatos_saida:
            col_saida = col
            print(f"  ✅✅ SAÍDA identificada (NOME EXATO): '{col}'")
            continue
    
    # PRIORIDADE 2: Se não encontrou pelos nomes exatos, busca por padrão (mais rigoroso)
    if col_entrada is None or col_saida is None:
        print(f"\n  Buscando por padrão (fallback)...")
        for col in todas_colunas:
            col_lower = col.lower().strip()
            
            # PULA se já foi identificada como quantidade
            if col == col_qtd_entrada or col == col_qtd_saida:
                continue
            
            # PULA se contém "qtd"
            if 'qtd' in col_lower:
                continue
            
            # Busca coluna de ENTRADA (valor monetário) - REGRA MUITO RÍGIDA
            if col_entrada is None:
                tem_entrada = 'entrada' in col_lower or 'receita' in col_lower or 'faturamento' in col_lower
                tem_r_dolar = 'r$' in col_lower or '$' in col_lower
                tem_nota = 'nota' in col_lower
                
                # REGRA ABSOLUTA: DEVE ter R$ E entrada E NÃO pode ter nota
                if tem_entrada and tem_r_dolar and not tem_nota:
                    col_entrada = col
                    print(f"  ✅ ENTRADA identificada (fallback): '{col}' (tem 'entrada' E 'R$' E sem 'nota')")
            
            # Busca coluna de SAÍDA (valor monetário) - REGRA MUITO RÍGIDA
            if col_saida is None:
                tem_saida = 'saida' in col_lower or 'saída' in col_lower or 'despesa' in col_lower
                tem_r_dolar = 'r$' in col_lower or '$' in col_lower
                tem_nota = 'nota' in col_lower
                
                # REGRA ABSOLUTA: DEVE ter R$ E saída E NÃO pode ter nota
                if tem_saida and tem_r_dolar and not tem_nota:
                    col_saida = col
                    print(f"  ✅ SAÍDA identificada (fallback): '{col}' (tem 'saída' E 'R$' E sem 'nota')")
    
    print("="*80 + "\n")
    
    # VALIDAÇÃO CRÍTICA FINAL
    print("\n" + "="*80)
    print("🔒 VALIDAÇÃO FINAL DAS COLUNAS IDENTIFICADAS")
    print("="*80)
    
    erro_encontrado = False
    
    if col_entrada:
        if 'qtd' in col_entrada.lower():
            print(f"❌ ERRO CRÍTICO: Coluna ENTRADA contém 'qtd': '{col_entrada}'")
            erro_encontrado = True
        else:
            print(f"✅ Coluna ENTRADA válida: '{col_entrada}'")
    
    if col_saida:
        if 'qtd' in col_saida.lower():
            print(f"❌ ERRO CRÍTICO: Coluna SAÍDA contém 'qtd': '{col_saida}'")
            erro_encontrado = True
        else:
            print(f"✅ Coluna SAÍDA válida: '{col_saida}'")
    
    if erro_encontrado:
        print("\n" + "="*80)
        print("🚨 ERRO: Colunas de quantidade foram identificadas como valores monetários!")
        print("="*80)
        raise ValueError(
            f"ERRO CRÍTICO: Colunas identificadas incorretamente!\n"
            f"Entrada: '{col_entrada}' (deveria ser coluna de VALOR, não quantidade)\n"
            f"Saída: '{col_saida}' (deveria ser coluna de VALOR, não quantidade)\n"
            f"Todas as colunas: {list(df.columns)}"
        )
    
    # Recupera nomes originais das colunas para debug
    col_mes_orig = colunas_originais.get(col_mes.lower().strip() if col_mes else "", col_mes or "N/A")
    col_entrada_orig = colunas_originais.get(col_entrada.lower().strip() if col_entrada else "", col_entrada or "N/A")
    col_saida_orig = colunas_originais.get(col_saida.lower().strip() if col_saida else "", col_saida or "N/A")
    col_qtd_entrada_orig = colunas_originais.get(col_qtd_entrada.lower().strip() if col_qtd_entrada else "", col_qtd_entrada or "N/A")
    col_qtd_saida_orig = colunas_originais.get(col_qtd_saida.lower().strip() if col_qtd_saida else "", col_qtd_saida or "N/A")
    
    # Mostra PRIMEIROS VALORES das colunas identificadas
    print("\n" + "="*80)
    print("🔍 VALORES DE EXEMPLO DAS COLUNAS IDENTIFICADAS")
    print("="*80)
    if col_entrada and len(df) > 0:
        exemplo_entrada = df[col_entrada].iloc[0] if len(df) > 0 else "N/A"
        print(f"  Coluna ENTRADA ('{col_entrada_orig}'): primeiro valor = {repr(exemplo_entrada)}")
    if col_saida and len(df) > 0:
        exemplo_saida = df[col_saida].iloc[0] if len(df) > 0 else "N/A"
        print(f"  Coluna SAÍDA ('{col_saida_orig}'): primeiro valor = {repr(exemplo_saida)}")
    if col_qtd_entrada and len(df) > 0:
        exemplo_qtd_entrada = df[col_qtd_entrada].iloc[0] if len(df) > 0 else "N/A"
        print(f"  Coluna QTD ENTRADA ('{col_qtd_entrada_orig}'): primeiro valor = {repr(exemplo_qtd_entrada)}")
    if col_qtd_saida and len(df) > 0:
        exemplo_qtd_saida = df[col_qtd_saida].iloc[0] if len(df) > 0 else "N/A"
        print(f"  Coluna QTD SAÍDA ('{col_qtd_saida_orig}'): primeiro valor = {repr(exemplo_qtd_saida)}")
    print("="*80 + "\n")
    
    # Adiciona ao debug_info (usa nomes originais)
    debug_info_colunas = {
        "excel_mes": col_mes_orig,
        "excel_entrada": col_entrada_orig,
        "excel_saida": col_saida_orig,
        "excel_qtd_entrada": col_qtd_entrada_orig,
        "excel_qtd_saida": col_qtd_saida_orig,
        "frontend_mes": "Mês/Ano",
        "frontend_entrada": "Entradas (R$)",
        "frontend_saida": "Saídas (R$)",
        "frontend_qtd_entrada": "Qtd. Notas Entrada",
        "frontend_qtd_saida": "Qtd. Notas Saída"
    }
    
    # Validação mínima
    if col_mes is None or col_entrada is None or col_saida is None:
        raise ValueError(
            "Planilha deve conter colunas: Mês/Período, Entradas/Receita, Saídas/Despesas. "
            f"Colunas encontradas: {', '.join(df.columns)}"
        )
    
    # Processa cada linha
    faturamento_mensal = []
    total_entrada = 0.0
    total_saida = 0.0
    
    # Armazena informações de debug
    debug_primeiras_linhas = []
    
    logger.info(f"=== DEBUG PROCESSAMENTO PLANILHA ===")
    logger.info(f"Todas as colunas encontradas: {list(df.columns)}")
    print("\n" + "="*70)
    print("🔍 COLUNAS IDENTIFICADAS")
    print("="*70)
    print(f"  ✓ Mês: {col_mes}")
    print(f"  ✓ Entrada (valor): {col_entrada}")
    print(f"  ✓ Saída (valor): {col_saida}")
    print(f"  ✓ Qtd. Entrada: {col_qtd_entrada}")
    print(f"  ✓ Qtd. Saída: {col_qtd_saida}")
    print(f"Total de linhas para processar: {len(df)}")
    print(f"Todas as colunas: {list(df.columns)}")
    print("="*70 + "\n")
    
    logger.info(f"Colunas identificadas:")
    logger.info(f"  - Mês: {col_mes}")
    logger.info(f"  - Entrada (valor): {col_entrada}")
    logger.info(f"  - Saída (valor): {col_saida}")
    logger.info(f"  - Qtd. Entrada: {col_qtd_entrada}")
    logger.info(f"  - Qtd. Saída: {col_qtd_saida}")
    logger.info(f"Total de linhas para processar: {len(df)}")
    
    # Mostra primeiras linhas para debug
    logger.info(f"Primeiras 3 linhas da planilha:")
    for i in range(min(3, len(df))):
        row = df.iloc[i]
        logger.info(f"  Linha {i}: Mês={row.get(col_mes, 'N/A')}, Entrada={row.get(col_entrada, 'N/A')}, Saída={row.get(col_saida, 'N/A')}")
    
    meses_nomes = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]
    
    print("\n" + "="*80)
    print("🚀 INICIANDO PROCESSAMENTO LINHA POR LINHA")
    print("="*80)
    print(f"Total de linhas no DataFrame: {len(df)}")
    print("="*80 + "\n")
    
    linha_numero = 0
    for idx, row in df.iterrows():
        linha_numero += 1
        try:
            print(f"\n{'='*80}")
            print(f"📝 LINHA {linha_numero} (índice do DataFrame: {idx})")
            print(f"{'='*80}")
            print(f"📋 DADOS BRUTOS DA LINHA:")
            for col in df.columns:
                print(f"   {col}: {repr(row[col])}")
            print(f"{'-'*80}")
            
            # Pula linhas completamente vazias
            if pd.isna(row[col_mes]) or str(row[col_mes]).strip() == '':
                print(f"⚠️  LINHA {linha_numero} VAZIA - PULANDO...")
                logger.debug(f"Linha {idx} vazia, pulando...")
                continue
            
            # Processa mês/período
            mes_str = str(row[col_mes]).strip()
            mes_formatado = mes_str
            
            # Pula se mês for inválido
            if mes_str.lower() in ['nan', 'none', '', 'observações', 'instruções']:
                print(f"⚠️  LINHA {linha_numero} COM MÊS INVÁLIDO '{mes_str}' - PULANDO...")
                logger.debug(f"Linha {idx} com mês inválido '{mes_str}', pulando...")
                continue
            
            print(f"✅ Mês identificado: '{mes_str}'")
            
            # Tenta extrair mês e ano
            mes_num = None
            ano = 2025
            
            # Formato: "Janeiro/2025" ou "01/2025" ou "2025-01" ou "01/01/2025" (DD/MM/YYYY)
            if '/' in mes_str:
                partes = mes_str.split('/')
                if len(partes) == 3:
                    # Formato DD/MM/YYYY (ex: "01/01/2025")
                    try:
                        dia = int(partes[0]) if partes[0].isdigit() else 1
                        mes_num = int(partes[1]) if partes[1].isdigit() else None
                        ano = int(partes[2]) if partes[2].isdigit() else 2025
                        logger.info(f"  Data identificada como DD/MM/YYYY: {dia}/{mes_num}/{ano}")
                    except:
                        mes_num = None
                elif len(partes) == 2:
                    # Formato: "Janeiro/2025" ou "01/2025"
                    if partes[0].isdigit():
                        mes_num = int(partes[0])
                        ano = int(partes[1]) if partes[1].isdigit() else 2025
                    else:
                        # Nome do mês
                        for i, nome in enumerate(meses_nomes, 1):
                            if nome.lower() in partes[0].lower():
                                mes_num = i
                                break
                        ano = int(partes[1]) if partes[1].isdigit() else 2025
            elif '-' in mes_str:
                # Formato: "2025-01"
                partes = mes_str.split('-')
                if len(partes) == 2:
                    ano = int(partes[0]) if partes[0].isdigit() else 2025
                    mes_num = int(partes[1]) if partes[1].isdigit() else None
            
            if mes_num is None:
                mes_num = (idx % 12) + 1  # Fallback: distribui pelos meses
            
            # Processa valores - trata formatos brasileiros e strings
            def converter_valor(valor, nome_coluna="valor"):
                """Converte valor para float, tratando formatos brasileiros."""
                if pd.isna(valor) or valor == "" or valor is None:
                    return 0.0
                
                # Se já é numérico
                if isinstance(valor, (int, float)):
                    return float(valor)
                
                # Converte para string e limpa
                valor_str = str(valor).strip()
                if not valor_str or valor_str.lower() in ['nan', 'none', '']:
                    return 0.0
                
                # Remove espaços e caracteres especiais (R$, $, etc)
                valor_limpo = valor_str.replace('r$', '').replace('$', '').replace(' ', '').strip()
                
                # Trata formato brasileiro: "100.000,50" -> 100000.50
                # Se tem vírgula, assume formato brasileiro (ponto = milhar, vírgula = decimal)
                if ',' in valor_limpo:
                    # Remove pontos (milhares) e substitui vírgula por ponto (decimal)
                    valor_limpo = valor_limpo.replace('.', '').replace(',', '.')
                # Se não tem vírgula mas tem ponto, pode ser formato americano ou brasileiro
                # Vamos assumir que se tem mais de 3 dígitos antes do ponto, é formato brasileiro
                elif '.' in valor_limpo:
                    partes_ponto = valor_limpo.split('.')
                    if len(partes_ponto) == 2 and len(partes_ponto[0]) > 3:
                        # Formato brasileiro sem vírgula: "100000.50" -> mantém como está
                        pass
                    # Caso contrário, já está no formato correto
                
                # Remove caracteres não numéricos exceto ponto e sinal negativo
                valor_limpo = ''.join(c for c in valor_limpo if c.isdigit() or c == '.' or c == '-')
                
                try:
                    return float(valor_limpo)
                except (ValueError, TypeError) as e:
                    logger.warning(f"Erro ao converter {nome_coluna} '{valor_str}': {e}")
                    return 0.0
            
            # VALIDAÇÃO CRÍTICA: Garante que está usando as colunas corretas
            if col_entrada and 'qtd' in col_entrada.lower():
                raise ValueError(f"ERRO CRÍTICO: Coluna de ENTRADA está incorreta! Está usando coluna de quantidade: '{col_entrada}'")
            if col_saida and 'qtd' in col_saida.lower():
                raise ValueError(f"ERRO CRÍTICO: Coluna de SAÍDA está incorreta! Está usando coluna de quantidade: '{col_saida}'")
            
            print(f"\n💰 CONVERTENDO VALORES:")
            print(f"   Coluna ENTRADA ('{col_entrada}'): {repr(row[col_entrada])} → ", end="")
            entrada = converter_valor(row[col_entrada], "entrada")
            print(f"R$ {entrada:,.2f}")
            
            print(f"   Coluna SAÍDA ('{col_saida}'): {repr(row[col_saida])} → ", end="")
            saida = converter_valor(row[col_saida], "saída")
            print(f"R$ {saida:,.2f}")
            
            qtd_entrada = int(converter_valor(row[col_qtd_entrada], "qtd_entrada")) if col_qtd_entrada and pd.notna(row[col_qtd_entrada]) else 0
            qtd_saida = int(converter_valor(row[col_qtd_saida], "qtd_saida")) if col_qtd_saida and pd.notna(row[col_qtd_saida]) else 0
            
            print(f"\n✅ RESULTADO DA LINHA {linha_numero}:")
            print(f"   Mês: {mes_str} → {mes_num}/{ano}")
            print(f"   Entrada: R$ {entrada:,.2f}")
            print(f"   Saída: R$ {saida:,.2f}")
            print(f"   Qtd. Entrada: {qtd_entrada}")
            print(f"   Qtd. Saída: {qtd_saida}")
            print(f"{'='*80}")
            
            # Log detalhado para debug - mostra qual coluna está sendo usada
            logger.info(f"═══════════════════════════════════════════════════")
            logger.info(f"PROCESSANDO LINHA {idx} (índice do DataFrame)")
            logger.info(f"  Mês original: '{mes_str}'")
            logger.info(f"  Valor BRUTO da coluna ENTRADA: {repr(row[col_entrada])} (tipo: {type(row[col_entrada])})")
            logger.info(f"  Valor BRUTO da coluna SAÍDA: {repr(row[col_saida])} (tipo: {type(row[col_saida])})")
            logger.info(f"  → Coluna ENTRADA usada: '{col_entrada}'")
            logger.info(f"  → Coluna SAÍDA usada: '{col_saida}'")
            logger.info(f"  → Entrada PROCESSADA: {entrada}")
            logger.info(f"  → Saída PROCESSADA: {saida}")
            logger.info(f"  → Mês numérico identificado: {mes_num}/{ano}")
            logger.info(f"═══════════════════════════════════════════════════")
            
            # Armazena primeiras linhas para debug
            if len(debug_primeiras_linhas) < 5:
                debug_primeiras_linhas.append({
                    "linha": int(idx),
                    "mes_original": mes_str,
                    "entrada_lida": str(row.get(col_entrada, 'N/A')),
                    "entrada_processada": entrada,
                    "saida_lida": str(row.get(col_saida, 'N/A')),
                    "saida_processada": saida,
                    "qtd_entrada": qtd_entrada,
                    "qtd_saida": qtd_saida
                })
            
            total_entrada += entrada
            total_saida += saida
            
            # Formata mês
            if mes_formatado == mes_str and mes_num:
                mes_formatado = f"{meses_nomes[mes_num - 1]}/{ano}"
            
            faturamento_mensal.append(FaturamentoMensal(
                mes=f"{ano}-{mes_num:02d}",
                mes_formatado=mes_formatado,
                total_entrada=round(entrada, 2),
                total_saida=round(saida, 2),
                quantidade_notas_entrada=qtd_entrada,
                quantidade_notas_saida=qtd_saida
            ))
            
        except Exception as e:
            logger.warning(f"Erro ao processar linha {idx}: {str(e)}")
            continue
    
    # Ordena por mês
    faturamento_mensal.sort(key=lambda x: x.mes)
    
    print("\n" + "="*70)
    print("✅ RESUMO DO PROCESSAMENTO")
    print("="*70)
    print(f"Total de meses processados: {len(faturamento_mensal)}")
    print(f"Total Entrada: R$ {total_entrada:,.2f}")
    print(f"Total Saída: R$ {total_saida:,.2f}")
    print(f"\n📊 DETALHE POR MÊS:")
    for fat in faturamento_mensal:
        print(f"  • {fat.mes_formatado}: Entrada=R$ {fat.total_entrada:,.2f}, Saída=R$ {fat.total_saida:,.2f}")
    print("="*70)
    print("✅ FIM DO PROCESSAMENTO\n")
    
    logger.info(f"═══════════════════════════════════════════════════")
    logger.info(f"RESUMO DO PROCESSAMENTO:")
    logger.info(f"  Total de meses processados: {len(faturamento_mensal)}")
    logger.info(f"  Total Entrada: R$ {total_entrada:,.2f}")
    logger.info(f"  Total Saída: R$ {total_saida:,.2f}")
    logger.info(f"  Detalhe por mês:")
    for fat in faturamento_mensal:
        logger.info(f"    {fat.mes_formatado}: Entrada={fat.total_entrada}, Saída={fat.total_saida}")
    logger.info(f"=== FIM DEBUG PROCESSAMENTO ===")
    logger.info(f"═══════════════════════════════════════════════════")
    
    # Calcula impactos
    impacto_cbs = total_saida * 0.12
    impacto_ibs = total_saida * 0.05
    impacto_previdencia = total_entrada * 0.02
    impacto_total = impacto_cbs + impacto_ibs - impacto_previdencia
    
    # Calcula período
    if faturamento_mensal:
        periodo_inicio = faturamento_mensal[0].mes
        periodo_fim = faturamento_mensal[-1].mes
        horizonte_temporal = f"Período: {faturamento_mensal[0].mes_formatado} até {faturamento_mensal[-1].mes_formatado}"
    else:
        periodo_inicio = "2025-01-01"
        periodo_fim = "2025-12-31"
        horizonte_temporal = "Ano 2025 - Período completo"
    
    # Formata CNPJ
    cnpj_limpo = ''.join(filter(str.isdigit, cnpj))
    cnpj_formatado = f"{cnpj_limpo[:2]}.{cnpj_limpo[2:5]}.{cnpj_limpo[5:8]}/{cnpj_limpo[8:12]}-{cnpj_limpo[12:]}" if len(cnpj_limpo) == 14 else cnpj
    
    resultado = {
        "nome_empresa": nome_empresa,
        "cnpj": cnpj_formatado,
        "total_entrada": round(total_entrada, 2),
        "total_saida": round(total_saida, 2),
        "impacto_cbs": round(impacto_cbs, 2),
        "impacto_ibs": round(impacto_ibs, 2),
        "impacto_previdencia": round(impacto_previdencia, 2),
        "impacto_total": round(impacto_total, 2),
        "periodo_inicio": periodo_inicio,
        "periodo_fim": periodo_fim,
        "periodo_mes": faturamento_mensal[-1].mes_formatado if faturamento_mensal else "Dezembro/2025",
        "horizonte_temporal": horizonte_temporal,
        "faturamento_mensal": faturamento_mensal,
        # Campos de debug (serão removidos antes de retornar)
        "_debug_col_mes": col_mes,
        "_debug_col_entrada": col_entrada,
        "_debug_col_saida": col_saida,
        "_debug_col_qtd_entrada": col_qtd_entrada,
        "_debug_col_qtd_saida": col_qtd_saida,
        "_debug_primeiras_linhas": debug_primeiras_linhas
    }
    
    return resultado

# -----------------------
# Endpoints
# -----------------------
@app.get("/")
async def root():
    return {"message": "API Dualtax ativa"}

@app.get("/download_template")
async def download_template():
    """Retorna template Excel estático para download."""
    import os
    from pathlib import Path
    
    # Tenta servir arquivo estático primeiro (mais confiável)
    # Tenta Excel primeiro, depois CSV
    template_path = Path("frontend/template_faturamento_dualtax.xlsx")
    template_csv_path = Path("frontend/template_faturamento_dualtax.csv")
    
    if template_path.exists():
        # Serve arquivo estático
        file_contents = template_path.read_bytes()
        return StreamingResponse(
            io.BytesIO(file_contents),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=template_faturamento_dualtax.xlsx"}
        )
    
    # Fallback: tenta CSV
    if template_csv_path.exists():
        file_contents = template_csv_path.read_bytes()
        return StreamingResponse(
            io.BytesIO(file_contents),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=template_faturamento_dualtax.csv"}
        )
    
    # Fallback: gera dinamicamente se openpyxl estiver disponível
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="Template não encontrado. Execute: python backend/gerar_template_estatico.py"
        )
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Faturamento Mensal"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="366092")
        instruction_font = Font(size=10, italic=True, color="666666")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = "📊 TEMPLATE DE FATURAMENTO MENSAL - DUALTAX"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # Instruções
        ws.merge_cells('A2:F2')
        ws['A2'] = "Preencha os dados abaixo com o faturamento mensal da sua empresa para calcular o impacto da Reforma Tributária"
        ws['A2'].font = instruction_font
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[3].height = 10
        
        # Cabeçalhos
        headers = ["Mês/Ano", "Entradas (R$)", "Saídas (R$)", "Qtd. Notas Entrada", "Qtd. Notas Saída", "Observações"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Largura das colunas
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 30
        
        # Dados de exemplo
        exemplos = [
            ["Janeiro/2025", 100000.00, 80000.00, 50, 45, "Exemplo"],
            ["Fevereiro/2025", 120000.00, 95000.00, 55, 50, ""],
            ["Março/2025", 110000.00, 85000.00, 52, 48, ""],
        ]
        
        for row_num, exemplo in enumerate(exemplos, 5):
            for col_num, valor in enumerate(exemplo, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = valor
                cell.border = border
                if col_num in [2, 3]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Salva em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=template_faturamento_dualtax.xlsx"}
        )
        
    except Exception as e:
        logger.error(f"Erro ao gerar template: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar template: {str(e)}")

def processar_notas_detalhadas(notas_detalhadas: List[dict]) -> List[Nota]:
    """
    Processa notas fiscais detalhadas da API e converte para formato interno.
    Calcula impostos considerando NCM, CNAE, alíquotas específicas, etc.
    """
    notas_processadas = []
    
    for nota_detalhada in notas_detalhadas:
        tipo = nota_detalhada.get('tipo', 'entrada')
        valor_total = nota_detalhada.get('valor_total', 0.0)
        
        # TODO: Calcular impostos por item considerando:
        # - NCM do produto
        # - CNAE da empresa
        # - Alíquotas específicas
        # - Regime tributário
        # - UF origem/destino
        
        notas_processadas.append(Nota(
            tipo=tipo,
            valor=float(valor_total)
        ))
    
    return notas_processadas

@app.get("/consultar_notas/{cnpj:path}", response_model=Resultado)
async def consultar_notas(cnpj: str):
    """Consulta notas fiscais por CNPJ."""
    try:
        logger.info(f"Recebida requisição para CNPJ: {cnpj}")
        
        # Remove formatação
        cnpj_limpo = ''.join(filter(str.isdigit, cnpj))
        logger.info(f"CNPJ limpo: {cnpj_limpo}")
        
        # Validação
        if len(cnpj_limpo) != 14:
            raise HTTPException(
                status_code=400,
                detail=f"CNPJ inválido. Deve conter 14 dígitos numéricos."
            )
        
        # Busca dados da empresa via API (APENAS DADOS REAIS - SEM FALLBACK)
        logger.info("Buscando dados da empresa via API...")
        dados_empresa = await buscar_dados_empresa_api(cnpj_limpo)
        logger.info(f"✅ Empresa obtida via API: {dados_empresa['nome_empresa']}")
        logger.info(f"   CNAE: {dados_empresa.get('cnae_principal', 'N/A')} - {dados_empresa.get('cnae_descricao', 'N/A')}")
        logger.info(f"   Fonte: {dados_empresa.get('fonte', 'N/A')}")
        
        # Busca notas fiscais via API (APENAS DADOS REAIS - SEM FALLBACK)
        logger.info("Buscando notas fiscais via API...")
        notas_detalhadas = await buscar_notas_fiscais_api(cnpj_limpo)
        
        # Notas fiscais - apenas dados reais
        notas = []
        if notas_detalhadas and len(notas_detalhadas) > 0:
            logger.info(f"✅ {len(notas_detalhadas)} notas encontradas via API")
            # Processa notas detalhadas (com NCM, alíquotas, etc.)
            notas = processar_notas_detalhadas(notas_detalhadas)
        else:
            logger.warning("⚠️ Nenhuma nota encontrada via API. Sistema retornará valores zerados.")
            # Não gera erro, mas retorna valores zerados (sem dados mockados)
            notas = []
        
        logger.info(f"Total de notas processadas: {len(notas)}")
        
        # Calcula totais (será zero se não houver notas)
        total_entrada = sum(n.valor for n in notas if n.tipo == 'entrada')
        total_saida = sum(n.valor for n in notas if n.tipo == 'saida')
        logger.info(f"Totais: Entrada={total_entrada}, Saída={total_saida}")
        
        # Se não houver notas, avisa mas não bloqueia
        if len(notas) == 0:
            logger.warning("⚠️ ATENÇÃO: Nenhuma nota fiscal encontrada. Valores serão zerados.")
        
        # Calcula impactos (será zero se não houver notas)
        impacto_cbs = total_saida * 0.12
        impacto_ibs = total_saida * 0.05
        impacto_previdencia = total_entrada * 0.02
        impacto_total = impacto_cbs + impacto_ibs - impacto_previdencia
        
        # Período
        logger.info("Calculando período...")
        periodo = calcular_periodo()
        logger.info(f"Período: {periodo}")
        
        # Faturamento mensal (APENAS DADOS REAIS - SEM MOCK)
        logger.info("Gerando faturamento mensal...")
        faturamento_mensal = gerar_faturamento_mensal(cnpj_limpo)
        if len(faturamento_mensal) > 0:
            logger.info(f"Faturamento mensal gerado: {len(faturamento_mensal)} meses")
        else:
            logger.info("Faturamento mensal não disponível (sem dados reais)")
        
        # Retorna resultado
        logger.info("Criando resultado...")
        
        # Adiciona avisos no debug_info se não houver dados
        debug_info = {
            "versao_processamento": API_VERSION,
            "fonte": dados_empresa.get("fonte", "API Externa"),
            "cnae_principal": dados_empresa.get("cnae_principal"),
            "cnae_descricao": dados_empresa.get("cnae_descricao"),
            "situacao": dados_empresa.get("situacao"),
            "regime_tributario": dados_empresa.get("regime_tributario"),
            "avisos": []
        }
        
        if len(notas) == 0:
            debug_info["avisos"].append("Nenhuma nota fiscal encontrada. Valores zerados.")
        
        if len(faturamento_mensal) == 0:
            debug_info["avisos"].append("Faturamento mensal não disponível. Sem dados reais.")
        
        resultado = Resultado(
            nome_empresa=dados_empresa["nome_empresa"],
            cnpj=dados_empresa["cnpj_formatado"],
            total_entrada=round(total_entrada, 2),
            total_saida=round(total_saida, 2),
            impacto_cbs=round(impacto_cbs, 2),
            impacto_ibs=round(impacto_ibs, 2),
            impacto_previdencia=round(impacto_previdencia, 2),
            impacto_total=round(impacto_total, 2),
            periodo_inicio=periodo["periodo_inicio"],
            periodo_fim=periodo["periodo_fim"],
            periodo_mes=periodo["periodo_mes"],
            horizonte_temporal=periodo["horizonte_temporal"],
            faturamento_mensal=faturamento_mensal,
            debug_info=debug_info
        )
        
        logger.info("Resultado criado com sucesso!")
        return resultado
        
    except HTTPException:
        raise
    except Exception as e:
        error_msg = f"Erro interno: {str(e)}\n{traceback.format_exc()}"
        logger.error(error_msg)
        raise HTTPException(status_code=500, detail=error_msg)

@app.post("/validar_planilha")
async def validar_planilha(file: UploadFile = File(...)):
    """Valida planilha antes do processamento - Consulta agente tributário."""
    try:
        logger.info(f"Validação de planilha: {file.filename}")
        
        file_ext = file.filename.split('.')[-1].lower()
        if file_ext not in ['xlsx', 'xls', 'csv']:
            raise HTTPException(
                status_code=400,
                detail="Formato não suportado. Use Excel (.xlsx, .xls) ou CSV (.csv)"
            )
        
        contents = await file.read()
        
        if file_ext == 'csv':
            df = pd.read_csv(io.BytesIO(contents), encoding='utf-8')
        else:
            # Tenta detectar automaticamente onde estão os cabeçalhos
            # O template tem: linha 1 (título), linha 2 (instruções), linha 3 (vazia), linha 4 (cabeçalhos)
            try:
                # Primeiro tenta ler com header na linha 4 (índice 3)
                df = pd.read_excel(io.BytesIO(contents), header=3)
                # Verifica se encontrou colunas válidas (não são "Unnamed")
                colunas_validas = [col for col in df.columns if not str(col).startswith('Unnamed') and not str(col).startswith('unnamed')]
                if len(colunas_validas) < 3:
                    # Tenta header=0 (primeira linha)
                    df_teste = pd.read_excel(io.BytesIO(contents), header=0)
                    colunas_validas_teste = [col for col in df_teste.columns if not str(col).startswith('Unnamed') and not str(col).startswith('unnamed')]
                    if len(colunas_validas_teste) >= 3:
                        df = df_teste
                    else:
                        # Tenta header=1, 2, 3...
                        for header_idx in range(4):
                            try:
                                df_teste = pd.read_excel(io.BytesIO(contents), header=header_idx)
                                colunas_validas_teste = [col for col in df_teste.columns if not str(col).startswith('Unnamed') and not str(col).startswith('unnamed')]
                                if len(colunas_validas_teste) >= 3:
                                    df = df_teste
                                    break
                            except:
                                continue
            except Exception as e:
                logger.warning(f"Erro ao detectar cabeçalhos automaticamente: {e}. Tentando leitura padrão...")
                # Fallback: lê normalmente
                df = pd.read_excel(io.BytesIO(contents))
        
        # Consulta agente tributário
        validacao = consultar_agente_tributario(df, "")
        
        return {
            "validacao": validacao.dict(),
            "colunas_encontradas": list(df.columns),
            "total_linhas": len(df)
        }
        
    except Exception as e:
        logger.error(f"Erro na validação: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao validar planilha: {str(e)}")

@app.post("/upload_planilha", response_model=Resultado)
async def upload_planilha(
    file: UploadFile = File(...),
    cnpj: str = Form(...),
    nome_empresa: str = Form(None)
):
    """Upload de planilha Excel/CSV com faturamento mensal."""
    try:
        logger.info(f"Recebido upload de arquivo: {file.filename}")
        
        # Valida tipo de arquivo
        if not file.filename:
            raise HTTPException(status_code=400, detail="Arquivo não fornecido")
        
        file_ext = file.filename.split('.')[-1].lower()
        if file_ext not in ['xlsx', 'xls', 'csv']:
            raise HTTPException(
                status_code=400,
                detail="Formato não suportado. Use Excel (.xlsx, .xls) ou CSV (.csv)"
            )
        
        # Lê o arquivo
        contents = await file.read()
        
        # Processa conforme o tipo
        if file_ext == 'csv':
            df = pd.read_csv(io.BytesIO(contents), encoding='utf-8')
        else:
            # CORREÇÃO: Lê o Excel pulando as 3 primeiras linhas (título, instruções, linha vazia)
            # Template tem: linha 1 (título), linha 2 (instruções), linha 3 (vazia), linha 4 (cabeçalhos)
            try:
                print("\n" + "="*70)
                print("📖 LENDO ARQUIVO EXCEL...")
                print("="*70)
                print("Tentando ler com header=3 (linha 4 do Excel)...")
                
                # Lê com header na linha 4 (índice 3)
                df = pd.read_excel(io.BytesIO(contents), header=3)
                print(f"✅ Excel lido! Colunas encontradas: {list(df.columns)}")
                
                # Verifica se encontrou colunas válidas
                colunas_validas = [col for col in df.columns if not str(col).startswith('Unnamed') and not str(col).startswith('unnamed')]
                print(f"Colunas válidas: {len(colunas_validas)} - {colunas_validas}")
                print("="*70)
                
                logger.info(f"Excel lido com header=3. Colunas: {list(df.columns)}")
                logger.info(f"Colunas válidas encontradas: {len(colunas_validas)} - {colunas_validas}")
                
                # Se não encontrou colunas válidas, tenta outras opções
                if len(colunas_validas) < 3:
                    logger.warning("Poucas colunas válidas encontradas. Tentando outras opções...")
                    # Tenta header=0 (primeira linha)
                    df_teste = pd.read_excel(io.BytesIO(contents), header=0)
                    colunas_validas_teste = [col for col in df_teste.columns if not str(col).startswith('Unnamed') and not str(col).startswith('unnamed')]
                    if len(colunas_validas_teste) >= 3:
                        df = df_teste
                        logger.info(f"Usando header=0. Colunas: {list(df.columns)}")
                    else:
                        # Tenta header=1, 2, 3...
                        for header_idx in range(1, 5):
                            try:
                                df_teste = pd.read_excel(io.BytesIO(contents), header=header_idx)
                                colunas_validas_teste = [col for col in df_teste.columns if not str(col).startswith('Unnamed') and not str(col).startswith('unnamed')]
                                if len(colunas_validas_teste) >= 3:
                                    df = df_teste
                                    logger.info(f"Usando header={header_idx}. Colunas: {list(df.columns)}")
                                    break
                            except Exception as e:
                                logger.warning(f"Erro ao tentar header={header_idx}: {e}")
                                continue
            except Exception as e:
                logger.error(f"Erro ao ler Excel: {e}. Tentando leitura padrão...")
                # Fallback: lê normalmente
                df = pd.read_excel(io.BytesIO(contents))
                logger.info(f"Excel lido com leitura padrão. Colunas: {list(df.columns)}")
        
        print("\n" + "="*70)
        print("📊 PLANILHA LIDA COMPLETA")
        print("="*70)
        print(f"Total de linhas: {len(df)}")
        print(f"Total de colunas: {len(df.columns)}")
        print(f"Colunas encontradas: {list(df.columns)}")
        print("="*70)
        print("\n📋 PRIMEIRAS 5 LINHAS BRUTAS DO DATAFRAME:")
        for i in range(min(5, len(df))):
            print(f"\n  ┌─ Linha {i} (índice {df.index[i]})")
            for col in df.columns:
                valor = df.iloc[i][col]
                print(f"  │ {col}: {repr(valor)} (tipo: {type(valor).__name__})")
            print(f"  └─")
        print("="*70 + "\n")
        
        logger.info(f"═══════════════════════════════════════════════════")
        logger.info(f"PLANILHA LIDA COMPLETA:")
        logger.info(f"  Total de linhas: {len(df)}")
        logger.info(f"  Total de colunas: {len(df.columns)}")
        logger.info(f"  Colunas: {list(df.columns)}")
        logger.info(f"═══════════════════════════════════════════════════")
        logger.info(f"PRIMEIRAS 5 LINHAS BRUTAS DO DATAFRAME:")
        for i in range(min(5, len(df))):
            logger.info(f"  Linha {i} (índice {df.index[i]}):")
            for col in df.columns:
                logger.info(f"    {col}: {repr(df.iloc[i][col])} (tipo: {type(df.iloc[i][col])})")
        logger.info(f"═══════════════════════════════════════════════════")
        
        # Valida CNPJ e nome
        if not cnpj:
            raise HTTPException(status_code=400, detail="CNPJ não fornecido")
        
        cnpj_limpo = ''.join(filter(str.isdigit, cnpj))
        if len(cnpj_limpo) != 14:
            raise HTTPException(
                status_code=400,
                detail="CNPJ inválido. Deve conter 14 dígitos numéricos."
            )
        
        nome_empresa = nome_empresa or "Empresa"
        
        # Consulta agente tributário antes de processar
        logger.info("Consultando agente tributário para validação...")
        validacao = consultar_agente_tributario(df, cnpj_limpo)
        
        if not validacao.valido:
            logger.warning(f"Agente tributário: {validacao.mensagem}")
            # Ainda processa, mas retorna aviso
        
        # Processa planilha com debug detalhado
        dados = processar_planilha_faturamento(df, cnpj_limpo, nome_empresa)
        
        # Adiciona informações de debug ao resultado
        mapeamento = dados.get("_debug_mapeamento", {})
        dados["debug_info"] = {
            "versao_processamento": API_VERSION,
            "fonte": "Upload de Planilha",  # Indica que veio de upload
            "mapeamento_colunas": {
                "excel_mes": mapeamento.get("excel_mes", dados.get("_debug_col_mes", "N/A")),
                "excel_entrada": mapeamento.get("excel_entrada", dados.get("_debug_col_entrada", "N/A")),
                "excel_saida": mapeamento.get("excel_saida", dados.get("_debug_col_saida", "N/A")),
                "excel_qtd_entrada": mapeamento.get("excel_qtd_entrada", dados.get("_debug_col_qtd_entrada", "N/A")),
                "excel_qtd_saida": mapeamento.get("excel_qtd_saida", dados.get("_debug_col_qtd_saida", "N/A")),
                "frontend_mes": "Mês/Ano",
                "frontend_entrada": "Entradas (R$)",
                "frontend_saida": "Saídas (R$)",
                "frontend_qtd_entrada": "Qtd. Notas Entrada",
                "frontend_qtd_saida": "Qtd. Notas Saída"
            },
            "colunas_identificadas": {
                "mes": dados.get("_debug_col_mes", "N/A"),
                "entrada": dados.get("_debug_col_entrada", "N/A"),
                "saida": dados.get("_debug_col_saida", "N/A"),
                "qtd_entrada": dados.get("_debug_col_qtd_entrada", "N/A"),
                "qtd_saida": dados.get("_debug_col_qtd_saida", "N/A")
            },
            "primeiras_linhas_processadas": dados.get("_debug_primeiras_linhas", [])
        }
        
        # Remove campos de debug temporários
        for key in list(dados.keys()):
            if key.startswith("_debug_"):
                del dados[key]
        
        # Adiciona validação ao resultado
        dados["validacao_tributaria"] = validacao.dict()
        
        # Cria resultado
        resultado = Resultado(**dados)
        
        logger.info(f"Planilha processada com sucesso: {len(dados['faturamento_mensal'])} meses")
        return resultado
        
    except HTTPException:
        raise
    except ValueError as e:
        logger.error(f"Erro de validação: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        error_msg = f"Erro ao processar planilha: {str(e)}\n{traceback.format_exc()}"
        logger.error(error_msg)
        raise HTTPException(status_code=500, detail=error_msg)

# -----------------------
# Executando localmente
# -----------------------
if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
