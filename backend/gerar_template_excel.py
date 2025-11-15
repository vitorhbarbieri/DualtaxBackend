"""
Script para gerar template Excel profissional e amigável
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def criar_template_excel(nome_arquivo="template_faturamento_dualtax.xlsx"):
    """Cria template Excel profissional para upload de faturamento."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento Mensal"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="366092")
    instruction_font = Font(size=10, italic=True, color="666666")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = "📊 TEMPLATE DE FATURAMENTO MENSAL - DUALTAX"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    # Instruções
    ws.merge_cells('A2:F2')
    ws['A2'] = "Preencha os dados abaixo com o faturamento mensal da sua empresa para calcular o impacto da Reforma Tributária"
    ws['A2'].font = instruction_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Espaço
    ws.row_dimensions[3].height = 10
    
    # Cabeçalhos
    headers = [
        "Mês/Ano",
        "Entradas (R$)",
        "Saídas (R$)",
        "Qtd. Notas Entrada",
        "Qtd. Notas Saída",
        "Observações"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Largura das colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 30
    
    # Dados de exemplo
    exemplos = [
        ["Janeiro/2025", 100000.00, 80000.00, 50, 45, "Exemplo de preenchimento"],
        ["Fevereiro/2025", 120000.00, 95000.00, 55, 50, ""],
        ["Março/2025", 110000.00, 85000.00, 52, 48, ""],
        ["Abril/2025", 130000.00, 100000.00, 60, 55, ""],
        ["Maio/2025", 125000.00, 98000.00, 58, 53, ""],
        ["Junho/2025", "", "", "", "", "Preencha com seus dados"],
    ]
    
    for row_num, exemplo in enumerate(exemplos, 5):
        for col_num, valor in enumerate(exemplo, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = valor
            cell.border = border
            if col_num in [2, 3]:  # Colunas de valores
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col_num in [4, 5]:  # Colunas de quantidade
                cell.alignment = center_align
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Notas importantes
    ws.merge_cells('A12:F15')
    nota_cell = ws['A12']
    nota_cell.value = """📌 INSTRUÇÕES IMPORTANTES:

• Mês/Ano: Use formato "Janeiro/2025" ou "01/2025" ou "2025-01"
• Entradas: Valor total de receitas/faturamento do mês (R$)
• Saídas: Valor total de despesas/compras do mês (R$)
• Qtd. Notas: Quantidade de notas fiscais (opcional, mas recomendado)
• Você pode adicionar mais linhas conforme necessário
• Remova as linhas de exemplo antes de fazer upload"""
    nota_cell.font = Font(size=9, color="333333")
    nota_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    nota_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    nota_cell.border = border
    
    # Rodapé
    ws.merge_cells('A17:F17')
    ws['A17'] = f"Template gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} - Dualtax - Reforma Tributária 2026"
    ws['A17'].font = Font(size=8, italic=True, color="999999")
    ws['A17'].alignment = center_align
    
    # Salva arquivo
    wb.save(nome_arquivo)
    print(f"✅ Template criado: {nome_arquivo}")
    return nome_arquivo

if __name__ == "__main__":
    criar_template_excel()

